# 开发者：苏畅
# 2023/5/15 0:24 于上海财经大学

from django.shortcuts import render, redirect
from app1 import models
from app1.utils.ModelForms import MessageModelForm
from openpyxl import load_workbook
from app1.utils.Logistic_Regression import Logistic_Regression, preProcessing
from django.contrib import messages
import pandas as pd

logreg = Logistic_Regression()


# 信息录入
def message(request):
    # 控制手动录入部分的输入入口
    if request.method == 'GET':
        info = MessageModelForm()
        return render(request, 'data_add.html', {'info': info})

    # 控制excel自动录入部分, 此处错误判断不全！！by sc
    elif request.method == 'POST' and request.FILES:
        try:
            excel_file = request.FILES['excel_file']  # 获取上传的 Excel 文件对象

            new_data = pd.read_excel(excel_file)  # 5.11 sc,gzj
            new_data = preProcessing(new_data)
            new_data_p = logreg.predict_proba(new_data.drop(['entrepre'], axis=1))[:, 0]

            wb = load_workbook(excel_file)  # 加载 Excel 文件
            sheet = wb.active  # 获取活动工作表

            rows = sheet.iter_rows(min_row=2)  # 迭代每一行  values_only=True
            # p_hat = logreg.predict(rows)

            for row, p in zip(rows, new_data_p):
                # row1 = preProcessing(row)
                # print(row)
                famer = MessageModelForm.Meta.model(
                    # name=row[0],
                    name=str(row[0].value),
                    # tel=row[1],
                    tel=str(row[1].value),
                    # entrepre=row[2],
                    entrepre=int(row[2].value),
                    # entretype=row[3],
                    famtype=int(row[3].value),
                    # relation=row[4],
                    relation=float(row[4].value),
                    # gender=row[5],
                    gender=int(row[5].value),
                    # age=row[6],
                    age=int(row[6].value),
                    # marriage=row[7],
                    marriage=int(row[7].value),
                    # religion=row[8],
                    religion=int(row[8].value),
                    # faminum=row[9],
                    faminum=int(row[9].value),
                    # maler=row[10],
                    maler=int(row[10].value),
                    # farmland=row[11],
                    farmland=float(row[11].value),
                    # income=row[12],
                    income=float(row[12].value),
                    # network=row[13],
                    network=int(row[13].value),
                    # mkt=row[14],
                    mkt=float(row[14].value),
                    # expectation=row[15],
                    expectation=int(row[15].value),
                    # handicraft=row[16],
                    handicraft=int(row[16].value),
                    # socistatu=row[17],
                    socistatu=int(row[17].value),
                    # hmnCapital=row[18],
                    hmnCapital=int(row[18].value),
                    # recptime=row[19],
                    recptime=row[19].value,
                    # recpplace=row[20],
                    recpplace=str(row[20].value),
                    # remark=row[21]
                    remark=str(row[21].value),
                    # id=row[22]
                    # entretype=int(row[22].value),  # 5.11 sc
                    possibility=p
                )
                famer.save()  # 保存对象到数据库中
        except:
            messages.error(request, "导入数据出错！请检查Excel表格的数据格式。")
            return redirect('/message/')

        return redirect('/table/')  # 上传完成后重定向到列表页面
    else:
        # 用户提交，进行数据校验
        form = MessageModelForm(data=request.POST)
        # 若合法，重定向到表格展示页面
        if form.is_valid():
            cleaned_data = pd.DataFrame([form.cleaned_data])
            p = logreg.predict(preProcessing(cleaned_data))
            form.instance.possibility = p
            form.save()
            return redirect('/table/')
        # 若不合法，将已输入数据和报错信息返回让用户修改，不清除
        return render(request, 'data_add.html', {'info': form})
