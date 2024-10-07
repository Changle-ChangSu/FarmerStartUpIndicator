# 开发者：苏畅
# 2023/6/12 1:48 于上海财经大学

from django.shortcuts import render, redirect
from app1 import models
from openpyxl import load_workbook
from app1.utils.ModelForms import EmpModelForm, EmpModelForm_with_pwd
from app1.utils.pagination import Pagination
from app1.utils.depart_search import Search
from django.contrib import messages
from app1.utils.encrypt import md5


def depart_list(request):
    """各部门职员列表"""
    queryset_dept = models.Department.objects.all()
    queryset_emp, filt, filt_string = Search(request).after_search()
    page_object = Pagination(request, queryset_emp, filt_string)

    context = {
        'queryset_dept': queryset_dept,
        'queryset_emp': page_object.page_queryset,  # 控制显示分完页的数据
        'filt': filt,  # 控制搜索条件回显
        'page_string': page_object.html(),  # 控制页码控件
        'page_size': page_object.page_size,  # 控制页面数据显示数量
        'total_page_count': page_object.total_page_count,  # 控制跳转输入框合法性
        'page': page_object.page,  # 控制执行删除操作后仍能返回到正确的页码上
        'div': page_object.div,  # 控制执行删除操作后仍能返回到正确的页码上
        'filt_string': filt_string
    }

    return render(request, 'emp_list.html', context)


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户提交的title数据
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门列表页面
    return redirect("/depart/list/?depart={}".format(title))


def depart_delete(request, title, filt_string):
    """删除部门"""
    models.Department.objects.filter(title=title).delete()

    page = request.GET.get('page')
    div = request.GET.get('div')
    page_size = request.GET.get('page_size')

    if div == '1' and page > '1':
        page = int(page) - 1

    return redirect('/depart/list/?page={}&page_size={}{}'.format(page, page_size, filt_string))


def emp_add(request):
    """添加部门员工"""
    # 控制手动录入部分的输入入口
    if request.method == 'GET':
        info = EmpModelForm_with_pwd()
        return render(request, 'emp_add.html', {'info': info})

        # 控制excel自动录入部分
    elif request.method == 'POST' and request.FILES:
        try:
            excel_file = request.FILES['excel_file']  # 获取上传的 Excel 文件对象

            wb = load_workbook(excel_file)  # 加载 Excel 文件
            sheet = wb.active  # 获取活动工作表

            rows = sheet.iter_rows(min_row=2)  # 迭代每一行  values_only=True

            emps = []
            for row in rows:
                dept = models.Department.objects.filter(title=str(row[4].value)).first()
                pwd = md5("123456")
                confirm_pwd = pwd
                ut = 1

                if not dept:
                    messages.error(request, "导入数据出错！请确保每个职员所在的部门存在。")
                    return redirect('/emp/add/')

                emp = EmpModelForm_with_pwd.Meta.model(
                    name=str(row[0].value),
                    tel=str(row[1].value),
                    email=str(row[2].value),
                    position=str(row[3].value),
                    department=dept,
                    password=pwd,
                    confirm_pwd=confirm_pwd,
                    usertype=ut,
                )
                emps.append(emp)
        except:
            messages.error(request, "导入数据出错！请检查Excel表格的数据格式。")
            return redirect('/emp/add/')

        for emp in emps:
            emp.save()  # 保存对象到数据库中
        return redirect('/depart/list/')  # 上传完成后重定向到列表页面

    else:
        # 用户提交，进行数据校验
        form = EmpModelForm_with_pwd(data=request.POST)
        # 若合法，重定向到表格展示页面
        if form.is_valid():
            form.save()
            return redirect('/depart/list/')
        # 若不合法，将已输入数据和报错信息返回让用户修改，不清除
        return render(request, 'data_add.html', {'info': form})


def emp_delete(request, nid, filt_string):
    """删除职员"""
    models.EmpInfo.objects.filter(id=nid).delete()

    page = request.GET.get('page')
    div = request.GET.get('div')
    page_size = request.GET.get('page_size')

    if div == '1' and page > '1':
        page = int(page) - 1

    return redirect('/depart/list/?page={}&page_size={}{}'.format(page, page_size, filt_string))


def emp_edit(request, nid, filt_string):
    """编辑职员"""
    page_now = request.GET.get('page')
    div = request.GET.get('div')
    page_size = request.GET.get('page_size')

    if div == '1' and page_now > '1':
        page_now = int(page_now) - 1

    row_info = models.EmpInfo.objects.filter(id=nid).first()
    if request.method == "GET":
        # 根据ID获取原始数据
        form = EmpModelForm(instance=row_info)
        return render(request, 'emp_edit.html', {'form': form, 'nid': nid})

    # 将用户新提交的数据更新到该行
    form = EmpModelForm(data=request.POST, instance=row_info)
    if form.is_valid():
        form.save()

        return redirect('/depart/list/?page={}&page_size={}{}'.format(page_now, page_size, filt_string))

    return render(request, 'emp_edit.html', {'form': form})


def emp_reset(request, nid, filt_string):
    """重置职员的密码"""
    emp = models.EmpInfo.objects.get(id=nid)
    emp.password = md5('123456')
    emp.save()

    page = request.GET.get('page')
    div = request.GET.get('div')
    page_size = request.GET.get('page_size')

    if div == '1' and page > '1':
        page = int(page) - 1

    return redirect('/depart/list/?page={}&page_size={}{}'.format(page, page_size, filt_string))